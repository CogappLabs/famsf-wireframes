"""Push the curator 12-Tier-1 medium mapping to the Medium audit workbook.

Seeds a new tab from `MediumFilterTaxonomy-12Tier1Terms.xlsx` (the curators'
re-organised medium taxonomy, 2026-07-03), whose Tier-1 axis is **material
group** (Prints, Ink & drawing media, Textiles & fiber, Paint & pigment,
Paper & parchment, Ceramic, Glass, Stone, Metal, Organic, Inorganic, Other) -
a different axis from the object-type "Facet review" tab already in the workbook.

Writes one tab, "12-Tier-1 medium map": one row per medium term with its full
Tier 1 / Tier 2 / Tier 3 path (forward-filled from the .xlsx block layout) plus
the leaf object count, an Approve? dropdown, and a Cleanup / note column carrying
the curators' flagged questions (the .xlsx cell comments + Notes column) forward
inline so nothing is lost. Native Sheets table (addTable) for typed columns,
banding and the dropdown, matching the existing tabs' house style.

Idempotent: drops the tab's existing table + rewrites on each run.

The source .xlsx is an uploaded Office file (Sheets API rejects it), so this
reads it locally with openpyxl - point SOURCE_XLSX at the download.

Run (from the wireframes repo root, after `gcloud auth application-default login`):

    uv run --with openpyxl --with google-api-python-client --with google-auth \
        --with-editable ~/git/cogapp-sheets \
        python scripts/push_tier1_medium_map_sheet.py
"""

from pathlib import Path

import openpyxl

from cogapp_sheets import Client

SHEET_ID = "14h4f-ZGnjjbBS6svQjRbdVjBKe7DBv9MifPLWy79FkI"
SOURCE_XLSX = Path.home() / "Downloads" / "MediumFilterTaxonomy-12Tier1Terms.xlsx"
XLSX_TAB = "12Tier1Terms"

TAB = "12-Tier-1 medium map"  # tab (sheet) title — free-form
TABLE_ID = "tier1_medium_map"
# Table name must be formula-safe: no leading digit, no spaces/hyphens. Distinct
# from the tab title above (which has neither restriction).
TABLE_NAME = "Tier1_medium_map"

HEADER = [
    "Tier 1 (Medium group)",
    "Tier 2",
    "Tier 3",
    "Term (leaf)",
    "Objects",
    "Approve?",
    "Cleanup / curator note",
]
APPROVE_OPTIONS = ["Yes", "Change Tier 1", "Change Tier 2/3", "Merge / normalise", "Drop"]

# .xlsx column indices (1-based, per the parsed layout).
C_TIER1, C_TIER2, C_TIER3 = 2, 3, 4
C_CNT1, C_CNT2, C_CNT3 = 5, 6, 7
C_NOTE = 8


def _cell(ws, r: int, c: int):
    v = ws.cell(r, c).value
    if v in (None, ""):
        return None
    if isinstance(v, str):
        v = v.strip()
        # Skip spreadsheet SUM formulas (roll-up rows have no leaf count).
        if v.startswith("="):
            return None
        return v or None
    return v


def _comment_text(ws, r: int, c: int) -> str | None:
    """The human line of a threaded cell comment, if any (strips the ID header)."""
    cell = ws.cell(r, c)
    if not cell.comment:
        return None
    lines = [ln.strip() for ln in cell.comment.text.splitlines()]
    # Threaded export looks like: "======", "ID#...", "Author (date)", "<text>".
    body = [ln for ln in lines if ln and not ln.startswith("=") and not ln.startswith("ID#")]
    if body and "(" in body[0] and body[0].endswith(")"):
        body = body[1:]  # drop the "Author (date)" line
    return " ".join(body) or None


def build_rows() -> list[list]:
    """Forward-fill the block layout into one full-path row per leaf term.

    Tier-1 rows carry only a group name; Tier-2 rows carry a name + (sometimes) a
    count; Tier-3 rows carry the leaf term + count. A leaf term is whichever of
    Tier 3 / Tier 2 actually names the term, so 2-tier groups (count on the
    Tier-2 row, no Tier-3) still emit a row.
    """
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=False)
    ws = wb[XLSX_TAB]

    def _t3_follows(row: int) -> bool:
        """True if a Tier-3 leaf appears before the next Tier-1/Tier-2 row.

        Distinguishes a Tier-2 *header* (children below) from a Tier-2 that is
        itself the leaf term (a 2-tier group, or a countless term like
        'Cellulose nitrate' with no rows under it).
        """
        for rr in range(row + 1, ws.max_row + 1):
            if _cell(ws, rr, C_TIER1) or _cell(ws, rr, C_TIER2):
                return False
            if _cell(ws, rr, C_TIER3):
                return True
        return False

    out: list[list] = [HEADER]
    tier1 = tier2 = ""
    for r in range(2, ws.max_row + 1):
        t1 = _cell(ws, r, C_TIER1)
        t2 = _cell(ws, r, C_TIER2)
        t3 = _cell(ws, r, C_TIER3)
        note = _cell(ws, r, C_NOTE)
        # Prefer the threaded comment (the curator's actual question) over the
        # terser Notes-column echo when both are present on the note cell.
        comment = _comment_text(ws, r, C_NOTE)
        note_out = comment or note or ""

        if t1:
            tier1 = t1
            tier2 = ""
            continue  # Tier-1 header row: sets context, not itself a leaf.
        if t3:  # Tier-3 leaf under the current Tier-1/Tier-2.
            cnt = _cell(ws, r, C_CNT3)
            out.append([tier1, tier2, t3, t3, cnt or "", "", note_out])
            continue
        if t2:
            if _t3_follows(r):  # Tier-2 header for the Tier-3 rows that follow.
                tier2 = t2
            else:  # 2-tier leaf: the term is on the Tier-2 row (count optional).
                cnt = _cell(ws, r, C_CNT2)
                out.append([tier1, t2, "", t2, cnt or "", "", note_out])
                tier2 = t2
    return out


def _tab_id(client: Client, title: str) -> int | None:
    meta = client.meta(fields="sheets(properties(sheetId,title))")
    for s in meta.get("sheets", []):
        if s["properties"]["title"] == title:
            return s["properties"]["sheetId"]
    return None


def ensure_tab(client: Client, title: str) -> int:
    """Return the tab's sheetId, creating it (and clearing it) if needed."""
    sid = _tab_id(client, title)
    if sid is not None:
        client.clear(f"'{title}'!A1:Z100000")
        return sid
    client.batch_update([{"addSheet": {"properties": {"title": title}}}])
    sid = _tab_id(client, title)
    assert sid is not None
    return sid


def drop_existing_tables(client: Client, sid: int) -> None:
    """Delete tables already on the tab so addTable won't collide on re-run."""
    meta = client.meta(fields="sheets(properties(sheetId),tables(tableId))")
    reqs = []
    for s in meta.get("sheets", []):
        if s["properties"]["sheetId"] != sid:
            continue
        for t in s.get("tables", []):
            reqs.append({"deleteTable": {"tableId": t["tableId"]}})
    if reqs:
        client.batch_update(reqs, tolerant=True)


def add_table_request(sid: int, n_rows: int) -> dict:
    """addTable request making the written range a native table with the dropdown."""
    col_props = []
    for i, label in enumerate(HEADER):
        prop: dict = {"columnIndex": i, "columnName": label, "columnType": "TEXT"}
        if label == "Objects":
            prop["columnType"] = "DOUBLE"
        if label == "Approve?":
            prop["columnType"] = "DROPDOWN"
            prop["dataValidationRule"] = {
                "condition": {
                    "type": "ONE_OF_LIST",
                    "values": [{"userEnteredValue": v} for v in APPROVE_OPTIONS],
                }
            }
        col_props.append(prop)
    return {
        "addTable": {
            "table": {
                "tableId": TABLE_ID,
                "name": TABLE_NAME,
                "range": {
                    "sheetId": sid,
                    "startRowIndex": 0,
                    "endRowIndex": n_rows + 1,  # + header
                    "startColumnIndex": 0,
                    "endColumnIndex": len(HEADER),
                },
                "columnProperties": col_props,
            }
        }
    }


def main() -> None:
    """Write the 12-Tier-1 medium-map tab to the Medium audit workbook."""
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source .xlsx not found: {SOURCE_XLSX}")

    print(f"Reading {SOURCE_XLSX.name} …", flush=True)
    rows = build_rows()
    n = len(rows) - 1
    n_notes = sum(1 for r in rows[1:] if r[-1])
    print(f"  {n} leaf rows, {n_notes} carry a cleanup note", flush=True)

    client = Client(SHEET_ID)
    print(f"Writing '{TAB}' …", flush=True)
    sid = ensure_tab(client, TAB)
    drop_existing_tables(client, sid)
    client.write(f"'{TAB}'!A1", rows)
    client.batch_update([add_table_request(sid, n)])

    print(
        f"Done. https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit",
        flush=True,
    )


if __name__ == "__main__":
    main()
